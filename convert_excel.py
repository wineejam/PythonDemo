#!/usr/bin/env python
# coding:utf-8
import os
import openpyxl
import pandas as pd


def insert_empty_line(f, nf, line_no=1, num=1):
    """
    :param f: 原文件名
    :param nf: 新文件名
    :param line_no:  在第几行插入
    :param num:  插入多少空行
    :return:
    """
    # 必须保证excel文件第一行为空，否则数据错乱,所以插入在第一行插入空行
    wb = openpyxl.load_workbook(f)
    st = wb.active
    nwb = openpyxl.Workbook()
    nst = nwb.active
    for row in range(1, st.max_row + 1):
        for cln in range(1, st.max_column + 1):
            if row < line_no:
                nst.cell(row=row, column=cln).value = st.cell(row=row, column=cln).value
            else:
                nst.cell(row=row + num, column=cln).value = st.cell(row=row, column=cln).value
    nwb.save(nf)

    if not os.path.exists(nf):
        raise FileNotFoundError
    os.remove(f)
    os.rename(nf, f)
    print("Excel表格预处理完成")


def convert_excel(file, newfile):
    """
    :param file: 原文件名
    :param newfile: 新文件名
    :return:
    """
    if not os.path.exists(file):
        raise FileNotFoundError
    # 判断第一行是否为空行，不是空行需要添加空行
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    try:
        first_row = " ".join([x.value for x in list(sheet.rows)[0]])
    except TypeError as e:
        # NoneType 也是代表为空行
        first_row = ""
    if first_row:
        print("检测到第一行不为空，需要插入空行，处理中...")
        insert_empty_line(file, newfile)

    df = pd.read_excel(file)
    tmp = {}
    n = -1
    for each in df.itertuples():
        idx = each.Index
        t = {each[1]: each[2], each[3]: each[4]}
        if idx % 4 == 0:
            n += 1
            tmp[n] = t
        else:
            tmp[n].update(t)

    df0 = pd.DataFrame().from_dict(tmp).T
    with pd.ExcelWriter(file, engine="openpyxl") as writer:
        # 没有下面这个语句的话excel表将完全被覆盖
        writer.book = openpyxl.load_workbook(file)
        df0.to_excel(writer, sheet_name="处理结果", index=None)
    print("转换完成，请打开原文件在[处理结果]中查看结果")


if __name__ == '__main__':
    filename = os.path.join(os.path.dirname(os.path.realpath(__file__)), "test.xlsx")
    new_filename = os.path.join(os.path.dirname(os.path.realpath(__file__)), "new_test.xlsx")
    convert_excel(filename, new_filename)
